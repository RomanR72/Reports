import os
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from copy import deepcopy
from datetime import datetime

# --- Конфигурация ---
INPUT_DIR = 'INPUT'
OUTPUT_DIR = 'OUTPUT'
TARGET_ORDER = [
    "Общее число собы�", "Общее количество", "Последние инциде",
    "Активы в инциден�", "Затронутые актив", "Распределение ал", 
    "Последние 10 алер�"
]

COLUMN_ORDER = {
    "Общее количество": ["metric", "value"],
    "Последние инциде": ["tenantID", "createdAt", "name", "priority", "status", "id", "severity", "tenantName"],
    "Активы в инциден�": ["tenantID", "tenantName", "name", "weight", "count", "id"],
    "Затронутые актив": ["frequency", "tenantID", "tenantName", "id", "displayName", "criticality"],
    "Распределение ал": ["value", "metric"],
    "Последние 10 алер�": ["tenantID", "correlationRuleName", "id", "tenantName", "severity", "name", "status", "firstSeen", "userName", "priority"]
}

MONTH_NAMES = {
    1: "января", 2: "февраля", 3: "марта", 4: "апреля",
    5: "мая", 6: "июня", 7: "июля", 8: "августа",
    9: "сентября", 10: "октября", 11: "ноября", 12: "декабря"
}

def ensure_directories_exist():
    """Создает необходимые каталоги, если они не существуют"""
    os.makedirs(INPUT_DIR, exist_ok=True)
    os.makedirs(OUTPUT_DIR, exist_ok=True)

def format_date_range(date_str):
    """Форматирует строку с датами в формат 'DD Month YYYY - DD Month YYYY'"""
    try:
        date_parts = date_str.split(" - ")
        if len(date_parts) != 2:
            return date_str
        
        start_date = datetime.strptime(date_parts[0].split('T')[0], "%Y-%m-%d")
        end_date = datetime.strptime(date_parts[1].split('T')[0], "%Y-%m-%d")
        
        formatted_start = f"{start_date.day} {MONTH_NAMES[start_date.month]} {start_date.year}"
        formatted_end = f"{end_date.day} {MONTH_NAMES[end_date.month]} {end_date.year}"
        
        return f"{formatted_start} - {formatted_end}"
    except Exception as e:
        print(f"Ошибка форматирования даты: {e}")
        return date_str

def count_data_rows(ws):
    """Подсчитывает количество строк с данными (исключая заголовки)"""
    count = 0
    for row in ws.iter_rows(min_row=3):
        if any(cell.value for cell in row):
            count += 1
    return count

def sum_metric_column(ws, sheet_name):
    """Суммирует значения в столбце metric и возвращает сумму"""
    metric_col_idx = None
    for idx, cell in enumerate(ws[2], 1):
        if cell.value == "metric":
            metric_col_idx = idx
            break
    
    if not metric_col_idx:
        print(f"Столбец 'metric' не найден в листе '{sheet_name}'")
        return None
    
    total = 0
    for row in ws.iter_rows(min_row=3, min_col=metric_col_idx, max_col=metric_col_idx):
        for cell in row:
            try:
                value = float(cell.value) if cell.value else 0
                total += value
            except (ValueError, TypeError):
                continue
    return total

def has_no_data(ws):
    """Проверяет наличие 'No Data' в первых 5 строках"""
    for row in ws.iter_rows(max_row=5):
        for cell in row:
            if cell.value and "No Data" in str(cell.value):
                return True
    return False

def calculate_column_width(ws, column_letter):
    """Вычисляет оптимальную ширину столбца"""
    max_length = 0
    for cell in ws[column_letter]:
        try:
            value = str(cell.value) if cell.value else ""
            length = len(value) * 1.1 + 2
            if length > max_length:
                max_length = length
        except:
            pass
    return min(max_length, 50)

def copy_first_row(source_ws, target_ws):
    """Копирует первую строку без изменений"""
    for col_idx, cell in enumerate(source_ws[1], 1):
        target_cell = target_ws.cell(row=1, column=col_idx, value=cell.value)
        if cell.has_style:
            target_cell.font = deepcopy(cell.font)
            target_cell.border = deepcopy(cell.border)
            target_cell.fill = deepcopy(cell.fill)
            target_cell.number_format = cell.number_format
            target_cell.alignment = deepcopy(cell.alignment)

def reorder_columns(ws, order):
    """Изменяет порядок столбцов с сохранением первой строки"""
    if ws.max_row < 2 or has_no_data(ws):
        return

    headers = [cell.value for cell in ws[2]]
    if not all(h in headers for h in order):
        print(f"Предупреждение: не все заголовки найдены в листе {ws.title}")
        return

    data = {h: [] for h in headers}
    styles = {h: [] for h in headers}
    
    for row in ws.iter_rows(min_row=3):
        for idx, cell in enumerate(row):
            if idx < len(headers):
                data[headers[idx]].append(cell.value)
                styles[headers[idx]].append({
                    'font': deepcopy(cell.font) if cell.has_style else None,
                    'border': deepcopy(cell.border) if cell.has_style else None,
                    'fill': deepcopy(cell.fill) if cell.has_style else None,
                    'number_format': cell.number_format,
                    'alignment': deepcopy(cell.alignment) if cell.has_style else None
                })

    for col in range(ws.max_column, 1, -1):
        ws.delete_cols(col)
    
    for col_idx, header in enumerate(order, 1):
        if col_idx > 1:
            ws.insert_cols(col_idx)
        
        ws.cell(row=2, column=col_idx, value=header)
        
        for row_idx, (value, style) in enumerate(zip(data[header], styles[header]), 3):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            if style['font']:
                for attr, val in style.items():
                    if val and hasattr(cell, attr):
                        setattr(cell, attr, val)

    for col_idx in range(1, len(order) + 1):
        col_letter = get_column_letter(col_idx)
        ws.column_dimensions[col_letter].width = calculate_column_width(ws, col_letter)

def rename_sheets(wb):
    """Переименовывает листы в формате 1-1, 1-2 и т.д."""
    for i, sheet in enumerate(wb.worksheets, 1):
        original_name = sheet.title
        new_name = f"1-{i}"
        try:
            sheet.title = new_name
            print(f"Переименован лист: '{original_name}' -> '{new_name}'")
        except Exception as e:
            print(f"Ошибка при переименовании листа '{original_name}': {str(e)}")

def process_workbook(input_path, output_path):
    """Обрабатывает один файл Excel"""
    try:
        wb_source = load_workbook(input_path)
        wb_dest = load_workbook(input_path)
        
        while len(wb_dest.worksheets) > 0:
            wb_dest.remove(wb_dest.worksheets[0])

        results = {
            "Общее количество": None,
            "Общее число собы�": None,
            "Затронутые актив": None,
            "Дата событий": None
        }

        # Сначала обрабатываем все листы
        for sheet_name in TARGET_ORDER:
            if sheet_name in wb_source.sheetnames:
                source_ws = wb_source[sheet_name]
                dest_ws = wb_dest.create_sheet(sheet_name)
                
                copy_first_row(source_ws, dest_ws)
                
                if sheet_name == "Общее число собы�":
                    date_cell = source_ws['B1']
                    if date_cell.value:
                        results["Дата событий"] = format_date_range(date_cell.value)
                    results[sheet_name] = sum_metric_column(source_ws, sheet_name)
                elif sheet_name == "Общее количество":
                    results[sheet_name] = sum_metric_column(source_ws, sheet_name)
                elif sheet_name == "Затронутые актив":
                    results[sheet_name] = count_data_rows(source_ws)
                
                if not has_no_data(source_ws):
                    print(f"Обработка листа: {sheet_name}")
                    
                    for row in source_ws.iter_rows(min_row=2):
                        for cell in row:
                            new_cell = dest_ws.cell(
                                row=cell.row, 
                                column=cell.column, 
                                value=cell.value
                            )
                            if cell.has_style:
                                new_cell.font = deepcopy(cell.font)
                                new_cell.border = deepcopy(cell.border)
                                new_cell.fill = deepcopy(cell.fill)
                                new_cell.number_format = cell.number_format
                                new_cell.alignment = deepcopy(cell.alignment)
                    
                    if sheet_name in COLUMN_ORDER:
                        reorder_columns(dest_ws, COLUMN_ORDER[sheet_name])
                    
                    if sheet_name == "Общее количество" and results[sheet_name] is not None:
                        dest_ws['H1'] = results[sheet_name]
                        dest_ws['H1'].number_format = '#,##0'
                        print(f"Сумма столбца 'metric': {results[sheet_name]:,.0f} (H1)")
                    
                    if sheet_name == "Общее число собы�":
                        if results[sheet_name] is not None:
                            dest_ws['G1'] = results[sheet_name]
                            dest_ws['G1'].number_format = '#,##0'
                            print(f"Сумма столбца 'metric': {results[sheet_name]:,.0f} (G1)")
                        
                        if results["Дата событий"] is not None:
                            dest_ws['H1'] = results["Дата событий"]
                            print(f"Форматированная дата: {results['Дата событий']} (H1)")
                    
                    if sheet_name == "Затронутые актив" and results[sheet_name] is not None:
                        dest_ws['H1'] = results[sheet_name]
                        print(f"Количество строк с данными: {results[sheet_name]} (H1)")
                else:
                    print(f"Пропуск преобразования листа '{sheet_name}' (содержит 'No Data')")
                    for row in source_ws.iter_rows(min_row=2):
                        for cell in row:
                            new_cell = dest_ws.cell(
                                row=cell.row, 
                                column=cell.column, 
                                value=cell.value
                            )
                            if cell.has_style:
                                new_cell.font = deepcopy(cell.font)
                                new_cell.border = deepcopy(cell.border)
                                new_cell.fill = deepcopy(cell.fill)
                                new_cell.number_format = cell.number_format
                                new_cell.alignment = deepcopy(cell.alignment)

        # Переименовываем листы после обработки
        rename_sheets(wb_dest)

        wb_dest.save(output_path)
        wb_source.close()
        wb_dest.close()
        print(f"Файл успешно обработан: {os.path.basename(output_path)}")
        return True
    except Exception as e:
        print(f"Ошибка при обработке файла {input_path}: {str(e)}")
        return False

def process_all_reports():
    """Обрабатывает все файлы в каталоге reports"""
    ensure_directories_exist()
    
    processed_count = 0
    error_count = 0
    
    for filename in os.listdir(INPUT_DIR):
        if filename.endswith('.xlsx'):
            input_path = os.path.join(INPUT_DIR, filename)
            output_path = os.path.join(OUTPUT_DIR, "processed_" + filename)
            
            print(f"\nНачата обработка файла: {filename}")
            if process_workbook(input_path, output_path):
                processed_count += 1
            else:
                error_count += 1
    
    print(f"\nОбработка завершена. Успешно: {processed_count}, с ошибками: {error_count}")

# Запуск обработки всех файлов
process_all_reports()